# Excel batch image insert utility v1.0

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
import sys


def get_titles():
    return 'xlsx image insert utility', 'xlsx image insert util v1.0'


def main(cell_w_str, cell_h_str, files):
    wb = Workbook()
    ws = wb.active

    # some constants, may be different from each languages and versions

    # cell native size multiplier (native cell size -> openpyxl cell size)
    # can be calibrated by checking applied cell size in excel
    cwnm, chnm = 24.3/23.6, 1

    # cell pixel size multiplier (openpyxl cell size -> pixel cell size)
    # can be calibrated by checking displayed cell size in pixels on 100% zoom in excel
    cwm, chm = 1800/180, 582/350

    # image pixel size multiplier (pixel image size -> openpyxl image size)
    # can be calibrated by checking displayed image size in pixels on 100% zoom in excel
    iwm, ihm = 4/5, 4/5

    # image pixel offset multiplier (pixel image offset -> openpyxl image offset)
    # can be calibrated by checking displayed image offset in pixels on 100% zoom in excel
    iwom, ihom = 4/5, 4/5

    cell_w_native = float(cell_w_str) * cwnm
    cell_h_native = float(cell_h_str) * chnm

    file_count = len(files)
    for i, file in enumerate(files):
        print('Image %02d/%02d' % (i + 1, file_count))

        #set target row, col, and their size
        row, col = i, 0
        ws.column_dimensions[get_column_letter(col + 1)].width = cell_w_native
        ws.row_dimensions[row + 1].height = cell_h_native
        cell_w, cell_h = cell_w_native * cwm, cell_h_native * chm

        #open image and get dimensions
        img = Image(file)
        img_w, img_h = img.width, img.height

        # calculate image dimensions to cell size
        cell_ratio = cell_w / cell_h
        image_ratio = img_w / img_h

        if image_ratio > cell_ratio:
            scale = cell_w / img_w
            x_offset = 0
            y_offset = (cell_h - img_h * scale)/2
        else:
            scale = cell_h / img_h
            x_offset = (cell_w - img_w * scale)/2
            y_offset = 0

        img_size = XDRPositiveSize2D(pixels_to_EMU(img_w * iwm * scale), pixels_to_EMU(img_h * ihm * scale))

        # insert image using OneCellAnchor (no stretching error)
        marker = AnchorMarker(col=col, colOff=pixels_to_EMU(x_offset * iwom), row=row, rowOff=pixels_to_EMU(y_offset * ihom))
        img.anchor = OneCellAnchor(_from=marker, ext=img_size)
        ws.add_image(img)

    # save file
    wb.save('images.xlsx')


if __name__ == '__main__':
    args = sys.argv
    del args[0]
    main(args[0], args[1], args[2:])
